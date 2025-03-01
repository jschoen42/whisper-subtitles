"""
    © Jürgen Schoenemeyer, 01.03.2025 15:26

    src/utils/excel.py

    PUBLIC:
     - check_excel_file_exists(filepath: Path | str) -> bool

     - read_excel_file(folderpath: Path | str, filename: str) -> None | Tuple[Workbook, float]
     - read_excel_worksheet(folderpath: str, filename: str, sheet_name: str) -> Tuple[Worksheet | None, float]
     - get_excel_worksheet(workbook: Workbook, sheet_name: str) -> Worksheet | None

     - get_cell_text(in_cell: Cell | MergedCell) -> str:
     - get_cell_value(in_cell: Cell | MergedCell, check_boolean: bool = True) -> bool | str

     - check_hidden_rows_columns(sheet: Worksheet) -> None
     - check_single_quotes(wb_name: str, cell_text: str, line_number: int, function_name: str) -> Tuple[bool, str]
     - check_double_quotes(wb_name: str, cell_text: str, line_number: int, function_name: str) -> Tuple[bool, str]

     - excel_date(date: datetime, time_zone_offset: tzoffset) -> float
     - convert_datetime( time_string: str ) -> int
     - seconds_to_timecode_excel(x: float) -> str
"""
from __future__ import annotations

import re
import unicodedata
import warnings

from datetime import datetime
from pathlib import Path
from typing import TYPE_CHECKING, Any, Tuple

from dateutil import parser
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from utils.file import check_file_exists, get_modification_timestamp
from utils.trace import Trace
from utils.util import format_timestamp

if TYPE_CHECKING:
    from dateutil.tz import tzoffset
    from openpyxl.cell.cell import Cell, MergedCell
    from openpyxl.workbook.workbook import Workbook

# UserWarning: Data Validation extension is not supported and will be removed
warnings.simplefilter("ignore")

def check_excel_file_exists(filepath: Path | str) -> bool:
    filepath = Path(filepath)

    folderpath = filepath.parent
    filename = filepath.name

    if Path(filename).suffix.lower() != ".xlsx":
        Trace.error(f"unkown extention '{filename}'")
        return False

    else:
        return check_file_exists(folderpath, filename)

def read_excel_file(folderpath: Path | str, filename: str) -> Tuple[Workbook | None, float]:
    filepath = Path(folderpath) / filename

    if check_excel_file_exists(filepath) is False:
        return None, 0

    try:
        workbook: Workbook = load_workbook(filename = filepath)
    except OSError as err:
        Trace.error(f"{err}")
        return None, 0

    return workbook, get_modification_timestamp(filepath)

def read_excel_worksheet(folderpath: Path | str, filename: str, sheet_name: str) -> Tuple[Worksheet | None, float]:
    filepath = Path(folderpath) / filename

    if check_excel_file_exists(filepath) is False:
        return None, 0.0

    try:
        workbook: Workbook = load_workbook(filename = filepath)
    except OSError as err:
        Trace.error(f"{err}")
        return None, 0.0

    try:
        sheet = workbook[sheet_name]
        assert isinstance(sheet, Worksheet)  # noqa: S101
    except AssertionError:
        Trace.error(f"'{sheet_name}' is not a Worksheet")
        return None, 0.0
    except KeyError as err:
        Trace.error(f"KeyError: {err}")
        return None, 0.0

    check_hidden_rows_columns(sheet)
    return sheet, get_modification_timestamp(filepath)

def get_excel_worksheet(workbook: Workbook, sheet_name: str) -> Worksheet | None:
    try:
        sheet = workbook[sheet_name]
        assert isinstance(sheet, Worksheet)  # noqa: S101
    except AssertionError:
        Trace.error(f"'{sheet_name}' is not a Worksheet")
        return None
    except KeyError as err:
        Trace.error(f"{err}")
        return None

    check_hidden_rows_columns(sheet)
    return sheet

# check if column(s) or row(s) are hidden

def check_hidden_rows_columns(sheet: Any) -> None:
    for key, value in sheet.column_dimensions.items():
        if value.hidden is True:
            if key != "A":
                Trace.warning( f"hidden column: {key}")

    for row_num, row_dimension in sheet.row_dimensions.items():
        if row_dimension.hidden is True:
            Trace.warning( f"hidden row: {row_num}")

######################################################################################
# get_cell_value with converting
#  - '\n'  -> '<br>
#  - '[-]' -> '&shy;'
#  - 'true' -> True, 'false' -> False, 'N/A' -> ''
#  - '[br]' -> <br>
#  - [b]...[/b] -> <b>...</b>, same with i, u, mark, sub, sup
#  - [nobr] -> <nobr>
#  - [hide] -> <hide> (custom)
######################################################################################

def get_cell_value(in_cell: Cell | MergedCell, check_boolean: bool = True) -> bool | str:
    if in_cell.value is None:
        return ""

    # data_type
    #   f: formular
    #   s: string
    #   n: numeric
    #   b: boolean
    #   n: null
    #   inlineStr: inline
    #   e: error
    #   str: formlar cached string

    if in_cell.data_type == "f":
        return f"formula not support: {in_cell.value}"

    txt = unicodedata.normalize("NFC", str(in_cell.value).rstrip())
    txt = txt.replace("\n", "<br>")
    txt = txt.replace("[-]", "&shy;")

    if check_boolean is True:
        if txt == "true":
            return True
        if txt == "false":
            return False

    if txt == "N/A":
        return ""

    return re.sub(r"\[(\/*)(br|b|i|u|mark|hide|nobr|sub|sup)\]", r"<\1\2>", txt)

def get_cell_text(in_cell: Cell | MergedCell) -> str:
    return str(get_cell_value(in_cell, check_boolean=False))


# if beginning or ending spaces are relevant in excel cells
#  -> warp text in single or double quotes " word"
# e.g. whisper DATEV dictionary (uses double quotes)

def check_single_quotes(wb_name: str, cell_text: str, line_number: int, function_name: str) -> Tuple[bool, str]:
    text = cell_text.strip()

    if text == "":
        return False, ""

    if text[:1] == "'" and text[-1:] == "'":
        return True, text[1:-1]
    else:
        Trace.error(f"{function_name} '{wb_name}': line {line_number} single quotes missing: |{text}|")
        return False, ""

def check_double_quotes(wb_name: str, cell_text: str, line_number: int, function_name: str) -> Tuple[bool, str]:
    text = cell_text.strip()

    if text == "":
        return False, ""

    if text[:1] == '"' and text[-1:] == '"':
        return False, text[1:-1]
    else:
        Trace.error(f"{function_name} '{wb_name}': line {line_number} doublequotes missing: |{text}|")
        return True, ""

# get excel excel (inner) date
# https://forum.openoffice.org/en/forum/viewtopic.php?t=108820#post_content529967
#
# start date: 30.12.1899 (+ 1 day)

def excel_date(date: datetime, time_zone_offset: tzoffset) -> float:
    day_in_seconds = 86400

    delta = date - datetime(1899, 12, 30, tzinfo=time_zone_offset)
    return delta.days + delta.seconds / day_in_seconds

def convert_datetime(time_string: str) -> int:
    my_time_string = parser.parse(time_string.replace("UTC", ""))

    my_timestamp = int(datetime.timestamp(my_time_string))

    # Trace.debug( f"{time_string} -> {my_time_string} => epoch: {my_timestamp}" )
    return my_timestamp

def seconds_to_timecode_excel(x: float) -> str:
    return format_timestamp(x, always_include_hours=False, decimal_marker=".")
