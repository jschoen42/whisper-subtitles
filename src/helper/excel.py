"""
    © Jürgen Schoenemeyer, 08.01.2025

    PUBLIC:
     - import_project_excel(pathname: str, filename: str, inType: str) -> Dict:
     - set_print_settings(ws: Any):
     - import_captions_excel(pathname: str, filename: str) -> List:
     - import_dictionary_excel(pathname: str, filename: str) -> Tuple[ Dict[str,list[str|int]], List[str], float ]:
     - update_dictionary_excel(pathname: str, filename: str, filename_update: str, column_name: str, data: Dict) -> None | bool:
     - import_ssml_rules_excel(pathname: str, filename: str) -> Dict:
     - import_hunspell_PreCheck_excel(pathname: str, filename: str) -> Tuple[list[str], List[str], List[List]]:
"""

import os

from typing import Any, Dict, List, Tuple
from pathlib import Path

import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle, Font, Alignment, PatternFill

# from openpyxl.worksheet.worksheet import Worksheet
# from openpyxl.workbook.workbook import Workbook
# from openpyxl.cell.cell import Cell

from utils.trace     import Trace
from utils.decorator import duration
from utils.file      import check_excel_file_exists, get_modification_timestamp
from utils.excel     import get_cell_text, check_quotes_error
from helper.captions import seconds_to_timecode_excel

######################################################################################
#
#   import_project_excel(pathname: str, filename: str, inType: str) -> Dict:
#
#   ./data/[project]/[project].xlsx
#
#   Zeile 2: Main Prompt
#
#   dann:
#     0. File:      Dateiname mp4
#     1. Sprecher   für sprecherspezifishce Auswertung
#     2: Variante:  wg. gemeinsame Dateiene bei LODAS + Lohn und Gehalt
#                    - 00_LODAS_LuG
#                    - 01_LODAS
#                    - 02_LuG
#     3: Intro:     für Test, wann Sprecher startet Intro ca. 12 sec, normal ca. 1 sec
#     4: no prompt: kein extra prompt für das video (obwohl Texte im Prompt stehen)
#     5: Prompt:    initial_prompt
#
#   return({
#       prompt: main_prompt,
#       parts:  [
#           speaker: Name
#           files [
#              file:     filename
#              folder:   variante
#              isIntro:  bool
#              prompt:   initial_prompt
#           ]
#       ]
#   ])
#
######################################################################################

def import_project_excel(pathname: Path | str, filename: str) -> bool | Dict:
    pathname = Path(pathname)
    filepath = pathname / filename

    if not check_excel_file_exists(filepath):
        Trace.error(f"file not found: {filepath}")
        return False

    try:
        wb = load_workbook(filename=filepath)
    except OSError as err:
        Trace.error(f"importExcel: {err}")
        return False

    try:
        sheet = wb.worksheets[0]  # wb["mediaList"]
    except KeyError as err:
        Trace.error(f"importExcel: {err}")
        return False

    filename = ""
    speaker = ""
    main_prompt = ""

    data: Dict[str, Any] = {}
    part: Dict[str, Dict[str, Any]] = {}
    speakers: List[str] = []

    for i in range(2, sheet.max_row + 1):
        row = sheet[i]

        if i == 2:
            if get_cell_text(row[4]).lower() == "x":
                main_prompt = ""
            else:
                main_prompt = get_cell_text(row[5])

        else:
            filename = get_cell_text(row[0])
            speaker  = get_cell_text(row[1])
            project  = get_cell_text(row[2])
            intro    = get_cell_text(row[3])
            noprompt = get_cell_text(row[4]).lower() == "x"
            prompt   = get_cell_text(row[5])

            if noprompt:
                prompt = ""

            if filename:
                if speaker not in speakers:
                    speakers.append(speaker)
                    part[speaker] = {
                        "speaker": speaker,
                        "files": []
                    }

                part[speaker]["files"].append({
                    "file": filename,
                    "folder": project,
                    "isIntro": intro.lower() == "x",
                    "prompt": prompt,
                })

                # print(f"media '{filename}', speaker '{speaker}', folder '{project}', prompt '{prompt}'")

    data["prompt"] = main_prompt
    data["parts"]  = []

    for _key, value in part.items():
        data["parts"].append(value)

    return data


######################################################################################
#
#  helper
#
######################################################################################

excel_column_format: Dict = {
    "source":  [40],
    "dest":    [40],
    "comment": [50],
}

def set_print_settings(ws: Any) -> None:
    # https://openpyxl.readthedocs.io/en/stable/print_settings.html

    def inch_to_mm(inch: float) -> float:
        return inch / 2.54

    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE

    ws.print_options.gridLines = True
    ws.print_options.horizontalCentered = True

    ws.print_title_rows = "1:1"  # the first row
    # ws.print_title_cols = "A:A" # the first column

    ws.page_margins.left   = inch_to_mm(1.5)
    ws.page_margins.right  = inch_to_mm(1.5)
    ws.page_margins.top    = inch_to_mm(1.5)
    ws.page_margins.bottom = inch_to_mm(1.75)

    ws.page_margins.footer = inch_to_mm(0.9)

    ws.oddFooter.right.text = "&P / &N"  # curr page / max page
    ws.oddFooter.left.text  = "&F / &A"  # filename / name worksheet

    # ws.page_setup.scale = 100

    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = False


######################################################################################
#
#   export_TextToSpeech_excel(pathname: str, filename: str, data: List) -> int:
#
#   data [
#       {'start': 6.08, 'end': 20.52, 'text': "Ich begrüße Sie ..."},
#       ...
#   ]
#
#   ./data/[project]/09_excel/[settings]/[video].xlsx
#
######################################################################################

excel_column_format_cc: Dict = {
    "start": [ 15, "head", "start", "time"],
    "end":   [ 15, "head", "end",   "time"],
    "X":     [  5, "head", "",      "mark"],
    "s-p":   [  5, "head", "",      "mark"],
    "text":  [100, "head", "text",  "text"],
    "pause": [ 10, "head", "pause", "mark"],
}

def export_TextToSpeech_excel(data: List, pathname: Path | str, filename: str) -> bool:
    pathname = Path(pathname)

    def patch_width(width: int) -> float:
        return width + 91 / 128

    def set_styles(wb: Any) -> None:
        style = NamedStyle(name="head")
        style.font = Font(name="Open Sans Bold", color="00ffffff", size=10)
        style.fill = PatternFill(fgColor="004f81bd", fill_type="solid")
        style.alignment = Alignment(vertical="top", horizontal="center")
        wb.add_named_style(style)

        style = NamedStyle(name="time")
        style.font = Font(name="Segoe UI", color="00000000", size=11)
        # style.fill = PatternFill(fgColor="00ffffff", fill_type = "solid")
        style.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
        wb.add_named_style(style)

        style = NamedStyle(name="mark")
        style.font = Font(name="Segoe UI", color="00000000", size=11)
        # style.fill = PatternFill(fgColor="00ffffff", fill_type = "solid")
        style.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
        wb.add_named_style(style)

        style = NamedStyle(name="text")
        style.font = Font(name="Segoe UI", color="00000000", size=11)
        # style.fill = PatternFill(fgColor="00ffffff", fill_type = "solid")
        style.alignment = Alignment(vertical="center", wrapText=True)
        wb.add_named_style(style)

    wb = openpyxl.Workbook()
    set_styles(wb)

    ws = wb.worksheets[0]
    ws.title = "whisper transcription"
    ws.freeze_panes = ws["A2"]

    set_print_settings(ws)

    def append_row(line_number: int, styles: List, values: List) -> None:
        for i, value in enumerate(values):
            ws.cell(line_number, i + 1).style = styles[i]
            ws.cell(line_number, i + 1).value = value

    header_style = []
    header_text  = []
    body_style   = []
    for i, entry in enumerate(excel_column_format_cc):
        values = excel_column_format_cc[entry]
        width = values[0]
        ws.column_dimensions[get_column_letter(i + 1)].width = patch_width(width)

        header_style.append(values[1])
        header_text.append(values[2])
        body_style.append(values[3])

    append_row(1, header_style, header_text)

    last_end = True
    count = 0
    for i, entry in enumerate(data):
        start = seconds_to_timecode_excel(entry["start"])
        end = seconds_to_timecode_excel(entry["end"])
        if i == 0 and last_end:
            mark = "x"
        else:
            mark = ""

        text = entry["text"]

        last_end = text[-1] != "…"

        if entry["pause"] == -1:
            # count += 1
            # append_row(count+1, body_style, [start, end, mark, "", text, ""])
            pass
        else:
            count += 1
            if entry["pause"] == 0 and count != 1:
                append_row(count + 1, body_style, [start, end, mark, ">", text, 0])
            else:
                append_row(count + 1, body_style, [start, end, mark, "p", text, 0])

    if not pathname.is_dir():
        try:
            os.makedirs(pathname)
            Trace.update(f"makedir: {pathname}")
        except OSError as err:
            error_msg = str(err).split(":")[0]
            Trace.error(f"{error_msg}: {pathname}")

    dest_path = Path(pathname, filename)
    try:
        wb.save(filename=dest_path)
        Trace.result(f"'{dest_path}'")
        return True
    except OSError as err:
        Trace.error(f"{err}")
        return False

######################################################################################
#
#   import_captions_excel(pathname: str, filename: str) -> List:
#
#   ./data/[project]/11_excelPolly/[video].xlsx
#
#     1. start: Timecode Start
#     2. end:   Timecode Ende
#     3: mark:  x = Start neuer Take
#     4: s(entence), p(paragraph)
#     5: text
#     6: extrapause
#
#   return([
#       [id, text, type, pause],
#       ...
#   ])
#
######################################################################################

def import_captions_excel(pathname: Path | str, filename: str) -> None | List:
    pathname = Path(pathname)
    filepath = pathname / filename

    if not check_excel_file_exists(filepath):
        Trace.error(f"[import_captions_excel] file not found: {filepath}")
        return None

    try:
        wb = load_workbook(filename=filepath)
    except OSError as err:
        Trace.error(f"[import_captions_excel] importExcel: {err}")
        return None

    try:
        ws = wb.worksheets[0]  # wb["mediaList"]
    except KeyError as err:
        Trace.error(f"[import_captions_excel] importExcel: {err}")
        return None

    result = []
    curr_start = ""
    curr_end   = ""

    text = ""
    curr_text: List = []
    for i in range(2, ws.max_row + 1):
        row = ws[i]

        start  = get_cell_text(row[0])
        end    = get_cell_text(row[1])
        marked = get_cell_text(row[2]).lower() == "x" # start of one ssml
        type   = get_cell_text(row[3]).lower()
        text   = get_cell_text(row[4])
        pause  = get_cell_text(row[5])
        if pause == "":
            pause = "0"

        if type != "" and type not in "psn>":
            Trace.error(f"{filename} unknown type {type} (use 's' (sentence), 'p' (paragraph), 'n' (nothing) or '>' for append)")
            type = "p"

        if marked:
            if len(curr_text) > 0:
                result.append([curr_start, curr_end, curr_text])
                curr_text = []
            curr_start = start

        curr_text.append({
            "id":    start,
            "text":  text,
            "type":  type,
            "pause": int(pause),
        })

        curr_end = end

    if text != "":
        result.append([curr_start, curr_end, curr_text])

    return result

######################################################################################
#
#   import_dictionary_excel(pathname: str, filename: str) -> Dict:
#
#   ./data/_dictionary/Dictionary.xlsx
#
#     1. Orginal:    immer in Anführungszeichen
#     2. Korrektur:  immer in Anführungszeichen
#     3: Anmerkung
#
#   return(
#        { original:  [correction, wb_name, row], ... },
#        { sheet: mayRow, ... }
#  )
#
######################################################################################

@duration("Custom text replacements loaded")
def import_dictionary_excel(pathname: Path | str, filename: str) -> None | Tuple[ Dict[str,list[str|int]], List[str], float ]:
    pathname = Path(pathname)
    filepath = pathname / filename

    if not check_excel_file_exists(filepath):
        Trace.error(f"file not found: {filepath}")
        return None

    try:
        wb = load_workbook(filename=filepath)
    except OSError as err:
        Trace.error(f"importExcel: {err}")
        return None

    sheet_names: List = []
    result: Dict = {}
    for wb_name in wb.sheetnames:
        if wb_name[:1] == "-":
            continue

        sheet_names.append(wb_name)

        ws = wb[wb_name]
        for i in range(2, ws.max_row + 1):
            row = ws[i]

            error, original = check_quotes_error(wb_name, str(get_cell_text(row[0])), i, "import_dictionary_excel")
            if error or original == "":
                continue

            error, correction = check_quotes_error(wb_name, str(get_cell_text(row[1])), i, "import_dictionary_excel")
            if not error and correction == "":
                Trace.error(f"'{wb_name}': line {i} '{original}' correction empty")
                continue

            if original == correction:
                Trace.error(f"'{wb_name}': line {i} '{original}' original == correction")
                continue

            if original in result:
                Trace.error(f"'{wb_name}': line {i} '{original}' double entries => {result[original]} / {correction}")
                continue

            result[original] = [correction, wb_name, i]

    Trace.result(f"{filename}: {len(result)} entries")

    return result, sheet_names, get_modification_timestamp(filepath)

######################################################################################
#
#   update_dictionary_excel(pathname: str, filename: str, data: Dict):
#
#   ./data/_dictionary/Dictionary.xlsx

def update_dictionary_excel(pathname: Path | str, filename: str, filename_update: str, column_name: str, data: Dict) -> None | bool:
    pathname = Path(pathname)
    source = pathname / filename

    def set_styles(wb: Any) -> None:
        style = NamedStyle(name="used")
        style.font = Font(name="Consolas", color="00000000", size=10)
        style.alignment = Alignment(vertical="top", horizontal="center")

        try:
            wb.add_named_style(style)
        except Exception as err:
            Trace.error(f"{err}")

    if not check_excel_file_exists(source):
        Trace.error(f"file not found: {source}")
        return None

    try:
        wb = load_workbook(filename=source)
    except OSError as err:
        Trace.error(f"importExcel: {err}")
        return None

    set_styles(wb)

    for wb_name in wb.sheetnames:
        if wb_name[:1] == "-":
            continue

        if wb_name in data:
            data_info_sheet = data[wb_name]
        else:
            Trace.error(f"sheet '{wb_name}' missing in update info")
            continue

        ws = wb[wb_name]

        row = -1
        for i in range(0, ws.max_column):
            if get_cell_text(ws[1][i]) == column_name:
                row = i
                break

        if row < 0:
            Trace.error(f"sheet '{wb_name}' - column '{column_name}' not found")
            continue

        for i in range(3, ws.max_row + 1):
            row_cells = ws[i]

            ws.cell(i, row + 1).value = ""
            # ws.cell(i, row+1).style = "Normal"

            if get_cell_text(row_cells[0]) != "":
                if str(i) in data_info_sheet:
                    ws.cell(i, row + 1).value = data_info_sheet[str(i)]
                else:
                    ws.cell(i, row + 1).value = 0

                if get_cell_text(row_cells[1]) != "":
                    ws.cell(i, row + 1).style = "used"

    dest_path = pathname / filename_update
    try:
        wb.save(filename=dest_path)
        Trace.result(f"'{dest_path}'")
        return True
    except OSError as err:
        Trace.error(f"{err}")
        return False

######################################################################################
#
#   import_ssml_rules_excel(pathname: str, filename: str) -> Dict:
#
#   ./data/_polly/AmazonPolly.xlsx
#     jedes Sheet für ein Template zustängig
#
#   Zeile 1
#     5: Template
#
#   ab Zeile 2
#     1. pre:    Kontext (immer in Anführungszeichen)
#     2. key:    Suchwort (immer in Anführungszeichen)
#     3. post:   Kontext (immer in Anführungszeichen)
#     4: value:
#     5: Kommentar
#
#   return({
#        "alias":  [ [pre, word, post, alias], ... ],
#        "phonem": [ [pre, word, post, phonem], ...]
#        "say_as": [ [pre, word, post, sayAs], ...]
#   })
#
#####################################################################################

def import_ssml_rules_excel(pathname: Path | str, filename: str) -> None | Dict:
    pathname = Path(pathname)
    filepath = pathname / filename

    if not check_excel_file_exists(filepath):
        Trace.error(f"file not found: {filepath}")
        return None

    try:
        wb = load_workbook(filename=filepath)
    except OSError as err:
        Trace.error(f"importExcel: {err}")
        return None

    def parse_ws(wb_name: str, ws: Any) -> Tuple[str, List]:
        rules: List = []

        _error, template = check_quotes_error(wb_name, ws["e1"].value, 1, "import_ssml_rules_excel")
        if template == "":
            return template, rules

        for i in range(2, ws.max_row + 1):
            row = ws[i]

            _error, key = check_quotes_error(wb_name, str(get_cell_text(row[1])), i, "import_ssml_rules_excel")
            if key != "":
                _error, pre   = check_quotes_error(wb_name, str(get_cell_text(row[0])), i, "import_ssml_rules_excel")
                _error, post  = check_quotes_error(wb_name, str(get_cell_text(row[2])), i, "import_ssml_rules_excel")
                _error, value = check_quotes_error(wb_name, str(get_cell_text(row[3])), i, "import_ssml_rules_excel")
                if value != "":
                    rules.append([pre, key, post, value])

        return template, rules

    result = {}
    trace_details = ""

    for wb_name in wb.sheetnames:
        if wb_name[:1] == "-":
            continue

        template, rules = parse_ws(wb_name, wb[wb_name])
        if template != "":
            result[wb_name] = {
                "template": template,
                "rules": rules,
            }

        trace_details += f" {wb_name}: {len(rules)},"

    Trace.result(f"{trace_details[:-1]}")
    return result


######################################################################################
#
#   import_hunspell_PreCheck_excel(pathname: str, filename: str) -> List:
#
#   ./data/_hunspell/PreCheck.xlsx
#
#     1. Text (immer in Anführungszeichen)
#     2: Anmerkung
#
#   return([
#       ["z.B.", "u.a", ...],
#       ["www.mydatev.de", ... , "xlsx", ... "Testgruber", ... ],
#       [ ["en", "bloc"], ["Société", "à", "responsabilité", "limitée"] ... ]
#   ])
#
#####################################################################################

def import_hunspell_PreCheck_excel(pathname: Path | str, filename: str) -> None | Tuple[list[str], List[str], List[List]]:
    pathname = Path(pathname)
    filepath = pathname / filename

    if not check_excel_file_exists(filepath):
        Trace.error(f"file not found: {filepath}")
        return None

    try:
        wb = load_workbook(filename=filepath)
    except OSError as err:
        Trace.error(f"importExcel: {err}")
        return None

    abbreviations_with_dot = []
    singles = []
    multiples = []
    for wb_name in wb.sheetnames:
        if wb_name[:1] == "-":
            continue

        ws = wb[wb_name]
        for i in range(2, ws.max_row + 1):
            row = ws[i]

            error, original = check_quotes_error(wb_name, str(get_cell_text(row[0])), i, "import_hunspell_PreCheck_excel")
            if not error and original != "":
                if wb_name == "specialDot":
                    abbreviations_with_dot.append(original)
                else:
                    multiple = original.replace("  ", " ").split(" ")
                    if len(multiple) > 1:
                        multiples.append(multiple)
                    elif original[-1] == ".":
                        abbreviations_with_dot.append(original)
                    else:
                        singles.append(original)

    Trace.result(f"{len(singles)} single, {len(multiples)} multiple, abbreviations_with_dot {len(abbreviations_with_dot)}")

    return abbreviations_with_dot, singles, multiples
