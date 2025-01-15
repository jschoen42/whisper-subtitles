"""
    © Jürgen Schoenemeyer, 11.01.2025

    PUBLIC:
     - import_project_excel(pathname: Path | str, filename: str) -> None | Dict[str, Any]:
     - export_TextToSpeech_excel(data: List[ColumnSubtitleInfo], pathname: Path | str, filename: str) -> bool:
     - import_captions_excel(pathname: Path | str, filename: str) -> None | List[ColumnSubtitleInfo]:
     - import_dictionary_excel(pathname: Path | str, filename: str) -> None | Tuple[ DictionaryResultDict, SheetNames, float ]:


     - update_dictionary_excel(pathname: str, filename: str, filename_update: str, column_name: str, data: Dict) -> None | bool:
     - import_ssml_rules_excel(pathname: str, filename: str) -> Dict:
     - import_hunspell_PreCheck_excel(pathname: str, filename: str) -> Tuple[list[str], List[str], List[List]]:

    PRIVATE:
     - def page_setup_print(ws: Any) -> None:

"""

from typing import Any, Dict, List, NamedTuple, Tuple, TypedDict, cast
from pathlib import Path
from enum import StrEnum, IntEnum

import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle, Font, Alignment, PatternFill

from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet._read_only import ReadOnlyWorksheet
from openpyxl.chartsheet.chartsheet import Chartsheet
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.cell.read_only import ReadOnlyCell

from utils.trace     import Trace
from utils.decorator import duration
from utils.file      import get_modification_timestamp, create_folder
from utils.excel     import check_excel_file_exists, get_cell_text, check_quotes_error
from utils.util      import export_json

from helper.captions import seconds_to_timecode_excel


######################################################################################
#
#   import_project_excel(pathname: Path | str, filename: str) -> None | Dict[str, Any]:
#
#   ./data/[project]/[project].xlsx
#
#   Zeile 2: Main Prompt
#
#   dann:
#     0. File:      Dateiname mp4
#     1. Sprecher   für sprecherspezifishce Auswertung
#     2: Variante:  wg. gemeinsame Dateiene bei LODAS + Lohn und Gehalt
#                    - 00_LODAS_LuG
#                    - 01_LODAS
#                    - 02_LuG
#     3: Intro:     für Test, wann Sprecher startet Intro ca. 12 sec, normal ca. 1 sec
#     4: no prompt: kein extra prompt für das video (obwohl Texte im Prompt stehen)
#     5: Prompt:    initial_prompt
#
#   return({
#       prompt: main_prompt,
#       parts:  [
#           speaker: Name
#           files [
#              file:     filename
#              folder:   variante
#              isIntro:  bool
#              prompt:   initial_prompt
#           ]
#       ]
#   ])
#
######################################################################################

def import_project_excel(pathname: Path | str, filename: str) -> None | Dict[str, Any]:
    pathname = Path(pathname)
    filepath = pathname / filename

    if not check_excel_file_exists(filepath):
        Trace.error(f"file not found: {filepath}")
        return None

    try:
        wb = load_workbook(filename=filepath)
    except OSError as err:
        Trace.error(f"importExcel: {err}")
        return None

    try:
        sheet: Worksheet | ReadOnlyWorksheet = wb.worksheets[0]
    except KeyError as err:
        Trace.error(f"importExcel: {err}")
        return None

    filename = ""
    speaker = ""
    main_prompt = ""

    data: Dict[str, Any] = {}
    part: Dict[str, Dict[str, Any]] = {}
    speakers: List[str] = []

    for i in range(2, sheet.max_row + 1): # type: ignore
        row = sheet[i]

        if i == 2:
            if get_cell_text(row[4]).lower() == "x":
                main_prompt = ""
            else:
                main_prompt = get_cell_text(row[5])

        else:
            filename = get_cell_text(row[0]) or ""
            speaker  = get_cell_text(row[1]) or ""
            project  = get_cell_text(row[2]) or ""
            intro    = get_cell_text(row[3]) or ""
            noprompt = get_cell_text(row[4]).lower() == "x"
            prompt   = get_cell_text(row[5]) or ""

            if noprompt:
                prompt = ""

            if filename:
                if speaker not in speakers:
                    speakers.append(speaker)
                    part[speaker] = {
                        "speaker": speaker,
                        "files": []
                    }

                part[speaker]["files"].append({
                    "file": filename,
                    "folder": project,
                    "isIntro": intro.lower() == "x",
                    "prompt": prompt,
                })

                # print(f"media '{filename}', speaker '{speaker}', folder '{project}', prompt '{prompt}'")

    data["prompt"] = main_prompt
    data["parts"]  = []

    for _key, value in part.items():
        data["parts"].append(value)

    return data

######################################################################################
#
#  worksheet: printing setting
#
######################################################################################

def page_setup_print(ws: Any) -> None:
    # https://openpyxl.readthedocs.io/en/stable/print_settings.html

    def inch_to_mm(inch: float) -> float:
        return inch / 2.54

    # Page Setup -> Page
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = False
    # ws.page_setup.scale = 100

    # Page Setup -> Margins
    ws.page_margins.left   = inch_to_mm(1.5)
    ws.page_margins.right  = inch_to_mm(1.5)
    ws.page_margins.top    = inch_to_mm(1.5)
    ws.page_margins.bottom = inch_to_mm(1.75)
    ws.page_margins.footer = inch_to_mm(0.9)
    ws.print_options.horizontalCentered = True

    # Page Setup -> Header/Footer
    ws.oddFooter.right.text = "&P / &N"  # curr page / max page
    ws.oddFooter.left.text  = "&F / &A"  # filename / name worksheet

    # Page Setup -> Sheet
    ws.print_options.gridLines = True
    ws.print_title_rows = "1:1"   # Rows to repeat at top
    # ws.print_title_cols = "A:A" # Columns to repeat al left

######################################################################################
#
#   export_TextToSpeech_excel(data: List[ColumnSubtitleInfo], pathname: Path | str, filename: str) -> bool:
#
#   ./data/[project]/09_excel/[settings]/[video].xlsx
#
######################################################################################

class ColumnSubtitleInfo(TypedDict):
    start: float
    end:   float
    text:  str
    pause: float

"""
  [
    {'start': 12.55, 'end': 18.59, 'text': "Herzlich willkommen zum Seminar 'Chancen der Digitalisierung im Rechnungswesen nutzen'.", 'pause': 0},
    {'start': 18.59, 'end': 20.43, 'text': '[pause: 1.84 sec]', 'pause': -1},
    {'start': 20.43, 'end': 23.51, 'text': 'Liebe Kolleginnen und Kollegen, ich darf mich kurz vorstellen.', 'pause': 1.84},
    {'start': 23.51, 'end': 23.85, 'text': '[pause: 0.34 sec]', 'pause': -1},
    ...
    # one paragraph:
    {'start': 100.11, 'end': 107.75, 'text': 'Im Kapitel 1 möchte ich Ihnen ganz gerne ein paar grundlegende Hinweise geben, wie Sie die Digitalisierung ins Rennen bringen.', 'pause': 0.72},
    {'start': 107.75, 'end': 121.27, 'text': 'Im Kapitel 2 befassen wir uns mit elektronischen Bankauszügen, den verschiedenen Varianten, wie wir an die Daten, die so oder so elektronisch bei der Bank vorhanden sind, optimal verarbeiten können.', 'pause': 0},
    {'start': 121.27, 'end': 125.17, 'text': 'Das Kapitel 3 betrifft die Lerndatei.', 'pause': 0},
    ...
  ]
"""

class SubtitleColumnFormat(NamedTuple): # substitute for not existing TypedList
    width:        int
    header_style: str
    header_text:  str
    body_style:   str

excel_format_subtitle: Dict[str, SubtitleColumnFormat] = {
    "start": SubtitleColumnFormat( 15, "head", "start", "time"), # 00:09.764
    "end":   SubtitleColumnFormat( 15, "head", "end",   "time"), # 00:18.464
    "X":     SubtitleColumnFormat(  5, "head", "",      "mark"), # x (start new block) | empty (continue block)
    "s-p":   SubtitleColumnFormat(  5, "head", "",      "mark"), # ssml p (paragraph) | s (sentence) | > (append)
    "text":  SubtitleColumnFormat(100, "head", "text",  "text"), # subtitle text
    "pause": SubtitleColumnFormat( 10, "head", "pause", "mark"), # extra pause in seconds (manual added later)
}

class Color(StrEnum):
    WHITE   = "00ffffff"
    BLUE    = "004f81bd"
    GREY    = "00a5a5a5"
    BLACK   = "00000000"
    ERROR   = "00FF8888" # red
    WARNING = "00fcd723" # yellow
    OK      = "0092d050" # green

class Fontname(StrEnum):
    HEAD = "Open Sans Bold"
    BODY = "Consolas"

class Fontsize(IntEnum):
    HEAD = 10
    BODY = 10 # 11



def export_TextToSpeech_excel(data: List[SubtitleColumnFormat], pathname: Path | str, filename: str) -> bool:
    pathname = Path(pathname)

    Trace.fatal(filename)
    export_json(pathname, filename, data, None)


    def patch_width(width: int) -> float:
        return width + 91 / 128

    # define cell style
    #  - head: 'head'
    #  - body: 'time', 'mark', 'text'

    def set_styles(wb: Any) -> None:
        style = NamedStyle(name="head")
        style.font = Font(name=Fontname.HEAD, color=Color.WHITE, size=Fontsize.HEAD)
        style.fill = PatternFill(fgColor=Color.BLUE, fill_type="solid")
        style.alignment = Alignment(vertical="top", horizontal="center")
        wb.add_named_style(style)

        style = NamedStyle(name="time")
        style.font = Font(name=Fontname.BODY, color=Color.BLACK, size=Fontsize.BODY)
        # style.fill = PatternFill(fgColor=Color.WHITE, fill_type = "solid")
        style.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
        wb.add_named_style(style)

        style = NamedStyle(name="mark")
        style.font = Font(name=Fontname.BODY, color=Color.BLACK, size=Fontsize.BODY)
        # style.fill = PatternFill(fgColor=Color.WHITE, fill_type = "solid")
        style.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
        wb.add_named_style(style)

        style = NamedStyle(name="text")
        style.font = Font(name=Fontname.BODY, color=Color.BLACK, size=Fontsize.BODY)
        # style.fill = PatternFill(fgColor=Color.WHITE, fill_type = "solid")
        style.alignment = Alignment(vertical="center", wrapText=True)
        wb.add_named_style(style)

    def append_row(ws, line_number: int, styles: List[str], values: List[str]) -> None:
        for i, value in enumerate(values):
            ws.cell(line_number, i + 1).style = styles[i]
            ws.cell(line_number, i + 1).value = value

    wb: Workbook = openpyxl.Workbook()
    set_styles(wb)

    ws: Worksheet | ReadOnlyWorksheet = wb.worksheets[0]
    if isinstance(ws, ReadOnlyWorksheet):
        Trace.error(f"ReadOnlyWorksheet {wb.worksheets[0]}")
        return False

    ws.title = "whisper transcription"
    ws.freeze_panes = ws["A2"]

    page_setup_print(ws)

    header_style: List[str] = []
    header_text: List[str] = []
    body_style: List[str] = []

    for i, entry in enumerate(excel_format_subtitle):
        values = excel_format_subtitle[entry]

        ws.column_dimensions[get_column_letter(i + 1)].width = patch_width(values[0])
        header_style.append(values[1])
        header_text.append(values[2])
        body_style.append(values[3])

    append_row(ws, 1, header_style, header_text)

    last_end = True
    count = 0

    for i in range(0, len(data)):
        subtitle_info = cast(ColumnSubtitleInfo, data[i])

        start = seconds_to_timecode_excel(subtitle_info["start"])
        end = seconds_to_timecode_excel(subtitle_info["end"])

        if i == 0 and last_end:
            mark = "x"
        else:
            mark = ""

        text = subtitle_info["text"]

        last_end = text[-1] != "…"

        if subtitle_info["pause"] == -1:
            # count += 1
            # append_row(ws, count+1, body_style, [start, end, mark, "", text, ""])
            pass
        else:
            count += 1
            if subtitle_info["pause"] == 0 and count != 1:
                append_row(ws, count + 1, body_style, [start, end, mark, ">", text, "0"])
            else:
                append_row(ws, count + 1, body_style, [start, end, mark, "p", text, "0"])

    create_folder(pathname)

    dest_path = Path(pathname, filename)
    try:
        wb.save(filename=dest_path)
        Trace.error(f"'{dest_path}'")
        return True
    except OSError as err:
        Trace.wait(f"{err}")
        return False

######################################################################################
#
#   import_captions_excel(pathname: str, filename: str) -> List:
#
#   ./data/[project]/11_excelPolly/[video].xlsx
#
#     1. start: Timecode Start
#     2. end:   Timecode Ende
#     3: mark:  x = Start neuer Take
#     4: s(entence), p(paragraph)
#     5: text
#     6: extrapause
#
#   return(

#      [
#       start,
#       end,
#       [id, text, type, pause],
#       ...
#        ]
#   ])
#
######################################################################################

class ColumnSubtitleExcel(TypedDict):
    start: float
    end:   float
    mark:  str
    type:  str
    text:  str
    pause: float

def import_captions_excel(pathname: Path | str, filename: str) -> None | List[List[Any]]:
    pathname = Path(pathname)
    filepath = pathname / filename

    if not check_excel_file_exists(filepath):
        Trace.error(f"[import_captions_excel] file not found: {filepath}")
        return None

    try:
        wb: Workbook = load_workbook(filename=filepath)
    except OSError as err:
        Trace.error(f"[import_captions_excel] importExcel: {err}")
        return None

    try:
        ws: Worksheet | ReadOnlyWorksheet = wb.worksheets[0]
    except KeyError as err:
        Trace.error(f"[import_captions_excel] importExcel: {err}")
        return None

    if ws.max_row is None:
        return None

    result: List[List[Any]] = []
    curr_start = ""
    curr_end   = ""

    text = ""
    curr_text: List[Any] = []

    for i in range(2, ws.max_row + 1):
        row = ws[i]

        start  = get_cell_text(row[0])
        end    = get_cell_text(row[1])
        marked = get_cell_text(row[2]).lower() == "x" # start of one ssml
        type   = get_cell_text(row[3]).lower()
        text   = get_cell_text(row[4])
        pause  = get_cell_text(row[5])
        if pause == "":
            pause = "0"

        if type != "" and type not in "psn>":
            Trace.error(f"{filename} unknown type {type} (use 's' (sentence), 'p' (paragraph), 'n' (nothing) or '>' for append)")
            type = "p"

        if marked:
            if len(curr_text) > 0:
                result.append([curr_start, curr_end, curr_text])
                curr_text = []
            curr_start = start

        curr_text.append({
            "id":    start,
            "text":  text,
            "type":  type,
            "pause": int(pause),
        })

        curr_end = end

    if text != "":
        result.append([curr_start, curr_end, curr_text])

    return result

######################################################################################
#
#   import_dictionary_excel(pathname: Path | str, filename: str) -> None | Tuple[ DictionaryResultDict, SheetNames, float ]:
#
#   ./data/_dictionary/Dictionary.xlsx
#
#     1. Orginal:    immer in Anführungszeichen
#     2. Korrektur:  immer in Anführungszeichen
#     3: Anmerkung
#
#   return(
#        { original:  [correction, wb_name, row], ... },
#        { sheet: mayRow, ... }
#  )
#
######################################################################################

"""
    Workbook
     - "normalize"
     - "allgmein"
     - "urls"
     - "Fallbeispiele"
     - "Spezielles"
     - "-sz" # starts with "-" -> will be ignored (not imported)

    Worksheet
     - 1: "original"
     - 2: "correction"
     - 3: "Anmerkung"
     - 4: "used v2" # reimport of used static with whisper large-v2
     - 5: "used v3" # reimport of used static with whisper large-v3

    return result, sheet_names, get_modification_timestamp(filepath)
                   List[str], List[]
"""

class DictionaryEntry(NamedTuple):
    correction: str
    sheet_name: str
    row: int

DictionaryResultDict = Dict[str, DictionaryEntry]
SheetNames = List[str]

@duration("Custom text replacements loaded")
def import_dictionary_excel(pathname: Path | str, filename: str) -> None | Tuple[ DictionaryResultDict, SheetNames, float ]:
    pathname = Path(pathname)
    filepath = pathname / filename

    if not check_excel_file_exists(filepath):
        Trace.error(f"file not found: {filepath}")
        return None

    try:
        wb: Workbook = load_workbook(filename=filepath)
    except OSError as err:
        Trace.error(f"importExcel: {err}")
        return None

    sheet_names: SheetNames = []
    result: DictionaryResultDict = {}

    for wb_name in wb.sheetnames:
        if wb_name[:1] == "-":
            continue

        sheet_names.append(wb_name)

        ws: Chartsheet | Worksheet | ReadOnlyWorksheet = wb[wb_name]
        if isinstance(ws, Chartsheet):
            Trace.error(f"Chartsheet {wb_name}")
            continue

        if ws.max_row is None:
            continue

        for i in range(2, ws.max_row + 1):
            row = ws[i]

            error, original = check_quotes_error(wb_name, str(get_cell_text(row[0])), i, "import_dictionary_excel")
            if error or original == "":
                continue

            error, correction = check_quotes_error(wb_name, str(get_cell_text(row[1])), i, "import_dictionary_excel")
            if not error and correction == "":
                Trace.error(f"'{wb_name}': line {i} '{original}' correction empty")
                continue

            if original == correction:
                Trace.error(f"'{wb_name}': line {i} '{original}' original == correction")
                continue

            if original in result:
                Trace.error(f"'{wb_name}': line {i} '{original}' double entries => {result[original]} / {correction}")
                continue

            result[original] = DictionaryEntry(correction, wb_name, i)

    modification_timestamp = get_modification_timestamp(filepath)

    Trace.result(f"{filename}: {len(result)} entries")
    return result, sheet_names, modification_timestamp


######################################################################################
#
#   update_dictionary_excel(pathname: str, filename: str, data: Dict):
#
#   ./data/_dictionary/Dictionary.xlsx

def update_dictionary_excel(pathname: Path | str, filename: str, filename_update: str, column_name: str, data: Dict[str, Any]) -> None | bool:
    pathname = Path(pathname)
    source = pathname / filename

    def set_styles(wb: Any) -> None:
        style = NamedStyle(name="used")
        style.font = Font(name=Fontname.BODY, color=Color.BLACK, size=Fontsize.HEAD)
        style.alignment = Alignment(vertical="top", horizontal="center")

        try:
            wb.add_named_style(style)
        except Exception as err:
            Trace.error(f"{err}")

    if not check_excel_file_exists(source):
        Trace.error(f"file not found: {source}")
        return None

    try:
        wb: Workbook = load_workbook(filename=source)
    except OSError as err:
        Trace.error(f"importExcel: {err}")
        return None

    set_styles(wb)

    for wb_name in wb.sheetnames:
        if wb_name[:1] == "-":
            continue

        if wb_name in data:
            data_info_sheet = data[wb_name]
        else:
            Trace.error(f"sheet '{wb_name}' missing in update info")
            continue

        ws: Chartsheet | Worksheet | ReadOnlyWorksheet = wb[wb_name]
        if isinstance(ws, Chartsheet) or isinstance(ws, ReadOnlyWorksheet):
            Trace.error(f"Chartsheet {wb_name}")
            continue

        row = -1
        for i in range(0, ws.max_column):
            if get_cell_text(ws[1][i]) == column_name:
                row = i
                break

        if row < 0:
            Trace.error(f"sheet '{wb_name}' - column '{column_name}' not found")
            continue

        for i in range(3, ws.max_row + 1):
            row_cells: MergedCell | ReadOnlyCell | Cell = ws[i]

            if isinstance(row_cells, Cell):
                clear_cell: MergedCell | ReadOnlyCell | Cell = ws.cell(i, row + 1)
                if isinstance(clear_cell, Cell):
                    clear_cell.value = ""

                if get_cell_text(row_cells[0]) != "": # type: ignore
                    mycell: MergedCell | ReadOnlyCell | Cell = ws.cell(i, row + 1)

                    if isinstance(mycell, Cell):
                        if str(i) in data_info_sheet:
                            mycell.value = data_info_sheet[str(i)]
                        else:
                            mycell.value = 0

                        if get_cell_text(row_cells[1]) != "": # type: ignore
                            mycell.style = "used"

    dest_path = pathname / filename_update
    try:
        wb.save(filename=dest_path)
        Trace.result(f"'{dest_path}'")
        return True
    except OSError as err:
        Trace.error(f"{err}")
        return False

######################################################################################
#
#   import_ssml_rules_excel(pathname: str, filename: str) -> Dict:
#
#   ./data/_polly/AmazonPolly.xlsx
#     jedes Sheet für ein Template zustängig
#
#   Zeile 1
#     5: Template
#
#   ab Zeile 2
#     1. pre:    Kontext (immer in Anführungszeichen)
#     2. key:    Suchwort (immer in Anführungszeichen)
#     3. post:   Kontext (immer in Anführungszeichen)
#     4: value:
#     5: Kommentar
#
#   return({
#        "alias":  [ [pre, word, post, alias], ... ],
#        "phonem": [ [pre, word, post, phonem], ...]
#        "say_as": [ [pre, word, post, sayAs], ...]
#   })
#
#####################################################################################

def import_ssml_rules_excel(pathname: Path | str, filename: str) -> None | Dict[str, Any]:
    pathname = Path(pathname)
    filepath = pathname / filename

    if not check_excel_file_exists(filepath):
        Trace.error(f"file not found: {filepath}")
        return None

    try:
        wb: Workbook = load_workbook(filename=filepath)
    except OSError as err:
        Trace.error(f"importExcel: {err}")
        return None

    def parse_ws(wb_name: str, ws: Any) -> Tuple[str, List[List[str]]]:
        rules: List[List[str]] = []

        _error, template = check_quotes_error(wb_name, ws["e1"].value, 1, "import_ssml_rules_excel")
        if template == "":
            return template, rules

        for i in range(2, ws.max_row + 1):
            row = ws[i]

            _error, key = check_quotes_error(wb_name, str(get_cell_text(row[1])), i, "import_ssml_rules_excel")
            if key != "":
                _error, pre   = check_quotes_error(wb_name, str(get_cell_text(row[0])), i, "import_ssml_rules_excel")
                _error, post  = check_quotes_error(wb_name, str(get_cell_text(row[2])), i, "import_ssml_rules_excel")
                _error, value = check_quotes_error(wb_name, str(get_cell_text(row[3])), i, "import_ssml_rules_excel")
                if value != "":
                    rules.append([pre, key, post, value])

        return template, rules

    result: Dict[str, Any] = {}
    trace_details = "import"

    for wb_name in wb.sheetnames:
        if wb_name[:1] == "-":
            continue

        template, rules = parse_ws(wb_name, wb[wb_name])
        if template != "":
            result[wb_name] = {
                "template": template,
                "rules": rules,
            }

        trace_details += f" {wb_name}: {len(rules)},"

    Trace.result(f"{trace_details[:-1]}")
    return result


######################################################################################
#
#   import_hunspell_PreCheck_excel(pathname: str, filename: str) -> List:
#
#   ./data/_hunspell/PreCheck.xlsx
#
#     1. Text (immer in Anführungszeichen)
#     2: Anmerkung
#
#   return([
#       ["z.B.", "u.a", ...],
#       ["www.mydatev.de", ... , "xlsx", ... "Testgruber", ... ],
#       [ ["en", "bloc"], ["Société", "à", "responsabilité", "limitée"] ... ]
#   ])
#
#####################################################################################

def import_hunspell_PreCheck_excel(pathname: Path | str, filename: str) -> None | Tuple[List[str], List[str], List[List[str]]]:
    pathname = Path(pathname)
    filepath = pathname / filename

    if not check_excel_file_exists(filepath):
        Trace.error(f"file not found: {filepath}")
        return None

    try:
        wb: Workbook = load_workbook(filename=filepath)
    except OSError as err:
        Trace.error(f"importExcel: {err}")
        return None

    abbreviations_with_dot: List[str] = []
    singles: List[str] = []
    multiples: List[List[str]] = []

    for wb_name in wb.sheetnames:
        if wb_name[:1] == "-":
            continue

        ws: Chartsheet | Worksheet | ReadOnlyWorksheet = wb[wb_name]
        if isinstance(ws, Chartsheet):
            Trace.error(f"Chartsheet {wb_name}")
            continue

        if ws.max_row is None:
            continue

        for i in range(2, ws.max_row + 1):
            row = ws[i]

            error, original = check_quotes_error(wb_name, str(get_cell_text(row[0])), i, "import_hunspell_PreCheck_excel")
            if not error and original != "":
                if wb_name == "specialDot":
                    abbreviations_with_dot.append(original)
                else:
                    multiple = original.replace("  ", " ").split(" ")
                    if len(multiple) > 1:
                        multiples.append(multiple)
                    elif original[-1] == ".":
                        abbreviations_with_dot.append(original)
                    else:
                        singles.append(original)

    Trace.result(f"{len(singles)} single, {len(multiples)} multiple, abbreviations_with_dot {len(abbreviations_with_dot)}")

    return abbreviations_with_dot, singles, multiples
