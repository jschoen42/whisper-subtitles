"""
    © Jürgen Schoenemeyer, 17.01.2025

    PUBLIC:
     - import_project_excel(pathname: Path | str, filename: str) -> None | Dict[str, Any]
     - import_dictionary_excel(pathname: Path | str, filename: str) -> None | Tuple[ DictionaryResultDict, SheetNames, float ]
     - import_hunspell_PreCheck_excel(pathname: str, filename: str) -> Tuple[list[str], List[str], List[List]]
     - import_captions_excel(pathname: Path | str, filename: str) -> None | List[ColumnSubtitleInfo]
     - import_ssml_rules_excel(pathname: Path | str, filename: str) -> None | Dict[str, Any]
"""

from typing import Any, Dict, List, NamedTuple, Tuple, TypedDict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from utils.trace     import Trace
from utils.decorator import duration
from utils.file      import get_modification_timestamp
from utils.excel     import check_excel_file_exists, get_cell_text, check_double_quotes

"""
    Excel: Project Infos -> SpeechToText

    Workbook
     - "Projektdetails"

    Worksheet
     - row 1: Main Prompt
     - column 1: File
     - column 2: Sprecher
     - column 3: Variante  (01_LODAS, 02_LuG, 03_LuG_LODAS)
     - column 4: Intro     (x)
     - column 5: no prompt (x)
     - column 6: Prompt

    Return
      - dict prompt: main_prompt
      - list parts: [speaker, files [filename, folder, isIntro, prompt]]
"""
@duration("import '{filename}'")
def import_project_excel(pathname: Path | str, filename: str) -> None | Dict[str, Any]:
    pathname = Path(pathname)
    filepath = pathname / filename

    if not check_excel_file_exists(filepath):
        Trace.error(f"file not found: {filepath}")
        return None

    try:
        wb = load_workbook(filename=filepath)
    except OSError as err:
        Trace.error(f"importExcel: {err}")
        return None

    try:
        sheet: Worksheet = wb.worksheets[0]
    except KeyError as err:
        Trace.error(f"importExcel: {err}")
        return None

    filename = ""
    speaker = ""
    main_prompt = ""

    result: Dict[str, Any] = {}
    part: Dict[str, Dict[str, Any]] = {}
    speakers: List[str] = []

    for i in range(2, sheet.max_row + 1):
        row = sheet[i]

        if i == 2:
            if get_cell_text(row[4]).lower() == "x":
                main_prompt = ""
            else:
                main_prompt = get_cell_text(row[5])

        else:
            filename = get_cell_text(row[0]) or ""
            speaker  = get_cell_text(row[1]) or ""
            project  = get_cell_text(row[2]) or ""
            intro    = get_cell_text(row[3]) or ""
            noprompt = get_cell_text(row[4]).lower() == "x"
            prompt   = get_cell_text(row[5]) or ""

            if noprompt:
                prompt = ""

            if filename:
                if speaker not in speakers:
                    speakers.append(speaker)
                    part[speaker] = {
                        "speaker": speaker,
                        "files": []
                    }

                part[speaker]["files"].append({
                    "file": filename,
                    "folder": project,
                    "isIntro": intro.lower() == "x",
                    "prompt": prompt,
                })

                # print(f"media '{filename}', speaker '{speaker}', folder '{project}', prompt '{prompt}'")

    result["prompt"] = main_prompt
    result["parts"]  = []

    for _key, value in part.items():
        result["parts"].append(value)

    return result

"""
    Excel: Dictionary

    Workbook
     - normalize
     - allgmein
     - urls
     - Fallbeispiele
     - Spezielles
     - -sz # starts with "-" -> will be ignored (not imported)

    Worksheet
     - 1: original word(s)
     - 2: correction word(s)
     - 3: comment
     - 4: used v2 # reimport of used static with whisper large-v2
     - 5: used v3 # reimport of used static with whisper large-v3

    Return
      - dict original: correction, sheet_name, row
      - list of sheet names
      - file timestamp
"""
class DictionaryEntry(NamedTuple):
    correction: str
    sheet_name: str
    row: int

DictionaryResultDict = Dict[str, DictionaryEntry]
SheetNames = List[str]

@duration("import '{filename}'")
def import_dictionary_excel(pathname: Path | str, filename: str) -> None | Tuple[ DictionaryResultDict, SheetNames, float ]:
    pathname = Path(pathname)
    filepath = pathname / filename

    if not check_excel_file_exists(filepath):
        Trace.error(f"file not found: {filepath}")
        return None

    try:
        wb: Workbook = load_workbook(filename=filepath)
    except OSError as err:
        Trace.error(f"importExcel: {err}")
        return None

    sheet_names: SheetNames = []
    result: DictionaryResultDict = {}

    for sheet_name in wb.sheetnames:
        if sheet_name[:1] == "-":
            continue

        sheet_names.append(sheet_name)

        ws: Worksheet = wb[sheet_name]

        # if ws.max_row is None:
        #     continue

        for i in range(2, ws.max_row + 1):
            row = ws[i]

            error, original = check_double_quotes(sheet_name, str(get_cell_text(row[0])), i, "import_dictionary_excel")
            if error or original == "":
                continue

            error, correction = check_double_quotes(sheet_name, str(get_cell_text(row[1])), i, "import_dictionary_excel")
            if not error and correction == "":
                Trace.error(f"'{sheet_name}': line {i} '{original}' correction empty")
                continue

            if original == correction:
                Trace.error(f"'{sheet_name}': line {i} '{original}' original == correction")
                continue

            if original in result:
                Trace.error(f"'{sheet_name}': row {i+1} '{original}' double entries => '{result[original].correction}' / '{correction}'")
                continue

            result[original] = DictionaryEntry(correction, sheet_name, i)

    modification_timestamp = get_modification_timestamp(filepath)

    Trace.result(f"{filename}: {len(result)} entries")
    return result, sheet_names, modification_timestamp


"""
    Excel: Hunspell PreCheck

    Workbook
     - Abkürzungen
     - Referenten
     - url
     - Musterfälle
     - sonstiges
     - multiple
     - specialDot (deprecated)

    Worksheet
     - 1: Text
     - 2: Anmerkung

    Return:
     - List: abbreviations_with_dot: z.B., Abs., ggf., ...
     - List: singles: www.mydatev.de, KOST1, Meier-Muster, BA-BEA ...
     - List: multiples: 'en block', 'Société à responsabilité limitée'
"""
@duration("import '{filename}'")
def import_hunspell_PreCheck_excel(pathname: Path | str, filename: str) -> None | Tuple[List[str], List[str], List[List[str]]]:
    pathname = Path(pathname)
    filepath = pathname / filename

    if not check_excel_file_exists(filepath):
        Trace.error(f"file not found: {filepath}")
        return None

    try:
        wb: Workbook = load_workbook(filename=filepath)
    except OSError as err:
        Trace.error(f"importExcel: {err}")
        return None

    abbreviations_with_dot: List[str] = []
    singles: List[str] = []
    multiples: List[List[str]] = []

    for sheet_name in wb.sheetnames:
        if sheet_name[:1] == "-":
            continue

        ws: Worksheet = wb[sheet_name]

        # if ws.max_row is None:
        #     continue

        for i in range(2, ws.max_row + 1):
            row = ws[i]

            error, original = check_double_quotes(sheet_name, str(get_cell_text(row[0])), i, "import_hunspell_PreCheck_excel")
            if not error and original != "":
                if sheet_name == "specialDot":
                    abbreviations_with_dot.append(original)
                else:
                    multiple = original.replace("  ", " ").split(" ")
                    if len(multiple) > 1:
                        multiples.append(multiple)
                    elif original[-1] == ".":
                        abbreviations_with_dot.append(original)
                    else:
                        singles.append(original)

    Trace.result(f"{len(singles)} single, {len(multiples)} multiple, abbreviations_with_dot {len(abbreviations_with_dot)}")

    return abbreviations_with_dot, singles, multiples


"""
    Excel: Captions -> TextToSpeech

    Workbook
     - "Projektdetails"

    Worksheet
     - 1: start timecode -> 00:00.500
     - 2: end timecode   -> 00:05.002
     - 3: new segement   -> x
     - 4: type
           -> 'p' Paragraph
           -> 's' Sentence
           -> 'n' nothing
           -> '>' merge
     - 5: text
     - 6: pause          -> 0

    Return
      - Segement
        [start_timecode, end_timecode, [id, text, type, pause], ...]
"""

class ColumnSubtitleExcel(TypedDict):
    start: float
    end:   float
    mark:  str
    type:  str
    text:  str
    pause: float

@duration("import '{filename}'")
def import_captions_excel(pathname: Path | str, filename: str) -> None | List[List[Any]]:
    pathname = Path(pathname)
    filepath = pathname / filename

    if not check_excel_file_exists(filepath):
        Trace.error(f"[import_captions_excel] file not found: {filepath}")
        return None

    try:
        wb: Workbook = load_workbook(filename=filepath)
    except OSError as err:
        Trace.error(f"[import_captions_excel] importExcel: {err}")
        return None

    try:
        ws: Worksheet = wb.worksheets[0]
    except KeyError as err:
        Trace.error(f"[import_captions_excel] importExcel: {err}")
        return None

    # if ws.max_row is None:
    #     return None

    result: List[List[Any]] = []
    curr_start = ""
    curr_end   = ""

    text = ""
    curr_text: List[Any] = []

    for i in range(2, ws.max_row + 1):
        row = ws[i]

        start  = get_cell_text(row[0])
        end    = get_cell_text(row[1])
        marked = get_cell_text(row[2]).lower() == "x" # start of one ssml
        type   = get_cell_text(row[3]).lower()
        text   = get_cell_text(row[4])
        pause  = get_cell_text(row[5])
        if pause == "":
            pause = "0"

        if type != "" and type not in "psn>":
            Trace.error(f"{filename} unknown type {type} (use 's' (sentence), 'p' (paragraph), 'n' (nothing) or '>' for append)")
            type = "p"

        if marked:
            if len(curr_text) > 0:
                result.append([curr_start, curr_end, curr_text])
                curr_text = []
            curr_start = start

        curr_text.append({
            "id":    start,
            "text":  text,
            "type":  type,
            "pause": int(pause),
        })

        curr_end = end

    if text != "":
        result.append([curr_start, curr_end, curr_text])

    return result


"""
    Excel: SSML Rules

    https://docs.aws.amazon.com/polly/latest/dg/supportedtags.html

    Rules: <pre>[key]<post> -> {value} -> template
     - <sub alias>: [i.V.m.] + {in Verbindung mit}        -> <sub alias="in Verbindung mit">i.V.m.</sub>
     - <phoneme>:   <selbst>[buchen]<des> + {bu:xən}      -> selbst<phoneme alphabet="ipa" ph="bu:xən">buchen</phoneme>des
     - <lang>:      <Liquidität >[as a Service] + {en-US} -> Liquidität<lang xml:lang="en-US">as a Service</lang>
     - <say-as>:    [BWA] + {characters}                  -> <say-as interpret-as="characters">BWA</say-as>

    Note
     - <phoneme> not available for AI voices

    Workbook
     - alias
     - phoneme
     - language
     - say-as

    Worksheet
     - 1: pre
     - 2: key
     - 3: post
     - 4: value
     - 5: <template>

    Worksheet - <template>
     - alias:    <sub alias="[value]">[key]</sub>
     - phoneme:  <phoneme alphabet="ipa" ph="[value]">[key]</phoneme>
     - language: <lang xml:lang="[value]">[key]</lang>
     - say_as:   <say-as interpret-as="[value]">[key]</say-as>

    Return
     - alias:    template: <template> / rules: [pre, key, post, value]
     - phoneme:  template: <template> / rules: [pre, key, post, value]
     - language: template: <template> / rules: [pre, key, post, value]
     - say_as:   template: <template> / rules: [pre, key, post, value]
"""

@duration("import '{filename}'")
def import_ssml_rules_excel(pathname: Path | str, filename: str) -> None | Dict[str, Any]:
    pathname = Path(pathname)
    filepath = pathname / filename

    if not check_excel_file_exists(filepath):
        Trace.error(f"file not found: {filepath}")
        return None

    try:
        wb: Workbook = load_workbook(filename=filepath)
    except OSError as err:
        Trace.error(f"importExcel: {err}")
        return None

    def parse_ws(sheet_name: str, ws: Any) -> Tuple[str, List[List[str]]]:
        rules: List[List[str]] = []

        _error, template = check_double_quotes(sheet_name, ws["e1"].value, 1, "import_ssml_rules_excel")
        if template == "":
            return template, rules

        for i in range(2, ws.max_row + 1):
            row = ws[i]

            _error, key = check_double_quotes(sheet_name, str(get_cell_text(row[1])), i, "import_ssml_rules_excel")
            if key != "":
                _error, pre   = check_double_quotes(sheet_name, str(get_cell_text(row[0])), i, "import_ssml_rules_excel")
                _error, post  = check_double_quotes(sheet_name, str(get_cell_text(row[2])), i, "import_ssml_rules_excel")
                _error, value = check_double_quotes(sheet_name, str(get_cell_text(row[3])), i, "import_ssml_rules_excel")
                if value != "":
                    rules.append([pre, key, post, value])

        return template, rules

    result: Dict[str, Any] = {}
    trace_details = "import"

    for sheet_name in wb.sheetnames:
        if sheet_name[:1] == "-":
            continue

        template, rules = parse_ws(sheet_name, wb[sheet_name])
        if template != "":
            result[sheet_name] = {
                "template": template,
                "rules": rules,
            }

        trace_details += f" {sheet_name}: {len(rules)},"

    Trace.result(f"{trace_details[:-1]}")
    return result
