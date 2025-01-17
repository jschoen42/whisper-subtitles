"""
    © Jürgen Schoenemeyer, 17.01.2025

    PUBLIC:
      - export_TextToSpeech_excel(data: List[ColumnSubtitleInfo], pathname: Path | str, filename: str) -> bool:

    PRIVATE:
     - def page_setup_print(ws: Any) -> None:

"""

from typing import Any, Dict, List, NamedTuple, TypedDict, cast
from pathlib import Path
from enum import StrEnum, IntEnum

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle, Font, Alignment, PatternFill

from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from utils.trace     import Trace
from utils.decorator import duration
from utils.file      import create_folder

from helper.captions import seconds_to_timecode_excel

######################################################################################
#
#   export_TextToSpeech_excel(data: List[ColumnSubtitleInfo], pathname: Path | str, filename: str) -> bool:
#
#   ./data/[project]/09_excel/[settings]/[video].xlsx
#
######################################################################################

class ColumnSubtitleInfo(TypedDict):
    start: float
    end:   float
    text:  str
    pause: float

"""
  [
    {'start': 12.55, 'end': 18.59, 'text': "Herzlich willkommen zum Seminar 'Chancen der Digitalisierung im Rechnungswesen nutzen'.", 'pause': 0},
    {'start': 18.59, 'end': 20.43, 'text': '[pause: 1.84 sec]', 'pause': -1},
    {'start': 20.43, 'end': 23.51, 'text': 'Liebe Kolleginnen und Kollegen, ich darf mich kurz vorstellen.', 'pause': 1.84},
    {'start': 23.51, 'end': 23.85, 'text': '[pause: 0.34 sec]', 'pause': -1},
    ...
    # one paragraph:
    {'start': 100.11, 'end': 107.75, 'text': 'Im Kapitel 1 möchte ich Ihnen ganz gerne ein paar grundlegende Hinweise geben, wie Sie die Digitalisierung ins Rennen bringen.', 'pause': 0.72},
    {'start': 107.75, 'end': 121.27, 'text': 'Im Kapitel 2 befassen wir uns mit elektronischen Bankauszügen, den verschiedenen Varianten, wie wir an die Daten, die so oder so elektronisch bei der Bank vorhanden sind, optimal verarbeiten können.', 'pause': 0},
    {'start': 121.27, 'end': 125.17, 'text': 'Das Kapitel 3 betrifft die Lerndatei.', 'pause': 0},
    ...
  ]
"""

class SubtitleColumnFormat(NamedTuple): # substitute for not existing TypedList
    width:        int
    header_style: str
    header_text:  str
    body_style:   str

excel_format_subtitle: Dict[str, SubtitleColumnFormat] = {
    "start": SubtitleColumnFormat( 15, "head", "start", "time"), # 00:09.764
    "end":   SubtitleColumnFormat( 15, "head", "end",   "time"), # 00:18.464
    "X":     SubtitleColumnFormat(  5, "head", "",      "mark"), # x (start new block) | empty (continue block)
    "s-p":   SubtitleColumnFormat(  5, "head", "",      "mark"), # ssml p (paragraph) | s (sentence) | > (append)
    "text":  SubtitleColumnFormat(100, "head", "text",  "text"), # subtitle text
    "pause": SubtitleColumnFormat( 10, "head", "pause", "mark"), # extra pause in seconds (manual added later)
}

class Color(StrEnum):
    WHITE   = "00ffffff"
    BLUE    = "004f81bd"
    GREY    = "00a5a5a5"
    BLACK   = "00000000"
    ERROR   = "00FF8888" # red
    WARNING = "00fcd723" # yellow
    OK      = "0092d050" # green

class Fontname(StrEnum):
    HEAD = "Open Sans Bold"
    BODY = "Segoe UI" # "Consolas"

class Fontsize(IntEnum):
    HEAD = 10
    BODY = 11

@duration("write '{filename}'")
def export_TextToSpeech_excel(data: List[SubtitleColumnFormat], pathname: Path | str, filename: str) -> bool:
    pathname = Path(pathname)

    # export_json(pathname, filename.replace(".xlsx", ".json"), data, None)

    def patch_width(width: int) -> float:
        return width + 91 / 128

    # define cell style
    #  - head: 'head'
    #  - body: 'time', 'mark', 'text'

    def set_styles(wb: Any) -> None:
        style = NamedStyle(name="head")
        style.font = Font(name=Fontname.HEAD, color=Color.WHITE, size=Fontsize.HEAD)
        style.fill = PatternFill(fgColor=Color.BLUE, fill_type="solid")
        style.alignment = Alignment(vertical="top", horizontal="center")
        wb.add_named_style(style)

        style = NamedStyle(name="time")
        style.font = Font(name=Fontname.BODY, color=Color.BLACK, size=Fontsize.BODY)
        # style.fill = PatternFill(fgColor=Color.WHITE, fill_type = "solid")
        style.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
        wb.add_named_style(style)

        style = NamedStyle(name="mark")
        style.font = Font(name=Fontname.BODY, color=Color.BLACK, size=Fontsize.BODY)
        # style.fill = PatternFill(fgColor=Color.WHITE, fill_type = "solid")
        style.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
        wb.add_named_style(style)

        style = NamedStyle(name="text")
        style.font = Font(name=Fontname.BODY, color=Color.BLACK, size=Fontsize.BODY)
        # style.fill = PatternFill(fgColor=Color.WHITE, fill_type = "solid")
        style.alignment = Alignment(vertical="center", wrapText=True)
        wb.add_named_style(style)

    def append_row(ws: Worksheet, line_number: int, styles: List[str], values: List[Any]) -> None:
        for i, value in enumerate(values):
            ws.cell(line_number, i + 1).style = styles[i]
            ws.cell(line_number, i + 1).value = value

    wb: Workbook = openpyxl.Workbook()
    set_styles(wb)

    ws: Worksheet = wb.worksheets[0]
    ws.title = "whisper transcription"
    ws.freeze_panes = ws["A2"]

    page_setup_print(ws)

    header_style: List[str] = []
    header_text: List[str] = []
    body_style: List[str] = []

    for i, entry in enumerate(excel_format_subtitle):
        values = excel_format_subtitle[entry]

        ws.column_dimensions[get_column_letter(i + 1)].width = patch_width(values[0])
        header_style.append(values[1])
        header_text.append(values[2])
        body_style.append(values[3])

    append_row(ws, 1, header_style, header_text)

    last_end = True
    count = 0

    for i in range(0, len(data)):
        subtitle_info = cast(ColumnSubtitleInfo, data[i])

        start = seconds_to_timecode_excel(subtitle_info["start"])
        end = seconds_to_timecode_excel(subtitle_info["end"])

        if i == 0 and last_end:
            mark = "x"
        else:
            mark = ""

        text = subtitle_info["text"]

        last_end = text[-1] != "…"

        if subtitle_info["pause"] == -1:
            # count += 1
            # append_row(ws, count+1, body_style, [start, end, mark, "", text, ""])
            pass
        else:
            count += 1
            if subtitle_info["pause"] == 0 and count != 1:
                append_row(ws, count + 1, body_style, [start, end, mark, ">", text, 0])
            else:
                append_row(ws, count + 1, body_style, [start, end, mark, "p", text, 0])

    create_folder(pathname)

    dest_path = Path(pathname, filename)
    try:
        wb.save(filename=dest_path)
        Trace.info(f"'{dest_path}'")
        return True
    except OSError as err:
        Trace.wait(f"{err}")
        return False

######################################################################################
#
#  worksheet: printing setting
#
######################################################################################

def page_setup_print(ws: Any) -> None:
    # https://openpyxl.readthedocs.io/en/stable/print_settings.html

    def inch_to_mm(inch: float) -> float:
        return inch / 2.54

    # Page Setup -> Page
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    # ws.page_setup.scale = 100
    ws.page_setup.fitToHeight = False
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # Page Setup -> Margins
    ws.page_margins.left   = inch_to_mm(1.5)
    ws.page_margins.right  = inch_to_mm(1.5)
    ws.page_margins.top    = inch_to_mm(1.5)
    ws.page_margins.bottom = inch_to_mm(1.75)
    ws.page_margins.footer = inch_to_mm(0.9)
    ws.print_options.horizontalCentered = True

    # Page Setup -> Header/Footer
    ws.oddFooter.right.text = "&P / &N"  # curr page / max page
    ws.oddFooter.left.text  = "&F / &A"  # filename / name worksheet

    # Page Setup -> Sheet
    ws.print_options.gridLines = True
    ws.print_title_rows = "1:1"   # Rows to repeat at top
    # ws.print_title_cols = "A:A" # Columns to repeat al left
