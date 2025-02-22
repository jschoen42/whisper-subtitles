"""
    © Jürgen Schoenemeyer, 22.02.2025

    src/helper/excel_update.py (openpyxl)

    PUBLIC:
      - update_dictionary_excel(pathname: Path | str, filename: str, filename_update: str, column_name: str, data: Dict[str, Any]) -> bool
"""
from __future__ import annotations

from enum import IntEnum, StrEnum
from pathlib import Path
from typing import TYPE_CHECKING, Any, Dict

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, NamedStyle

from utils.decorator import duration
from utils.excel import check_excel_file_exists, get_cell_text
from utils.trace import Trace

if TYPE_CHECKING:
    from openpyxl.workbook.workbook import Workbook
    from openpyxl.worksheet.worksheet import Worksheet

class Color(StrEnum):
    WHITE     = "ffffff"
    BLUE      = "4f81bd"
    GREY      = "a5a5a5"
    LIGHTGREY = "dadcdd"
    BLACK     = "000000"
    ERROR     = "FF8888" # red
    WARNING   = "fcd723" # yellow
    OK        = "92d050" # green

class Fontname(StrEnum):
    HEAD = "Open Sans Bold"
    BODY = "Consolas"

class Fontsize(IntEnum):
    HEAD = 10
    BODY = 11

@duration("update '{filename}' -> '{filename_update}'")
def update_dictionary_excel(pathname: Path | str, filename: str, filename_update: str, column_name: str, data: Dict[str, Any]) -> bool:
    pathname = Path(pathname)
    source = pathname / filename

    def set_styles(wb: Workbook) -> None:
        if "used" not in wb.style_names:
            style = NamedStyle(name="used")
            style.font = Font(name=Fontname.BODY, color=Color.BLACK, size=Fontsize.HEAD)
            style.alignment = Alignment(vertical="top", horizontal="center")
            wb.add_named_style(style)

    if not check_excel_file_exists(source):
        Trace.error(f"file not found: {source}")
        return False

    try:
        wb: Workbook= load_workbook(filename=source)
    except OSError as err:
        Trace.error(f"importExcel: {err}")
        return False

    set_styles(wb)

    for sheet_name in wb.sheetnames:
        if sheet_name[:1] == "-":
            continue

        if sheet_name in data:
            data_info_sheet = data[sheet_name]
        else:
            Trace.error(f"sheet '{sheet_name}' missing in update info")
            continue

        ws: Worksheet = wb[sheet_name]

        row = -1
        for i in range(ws.max_column):
            if get_cell_text(ws[1][i]) == column_name:
                row = i
                break

        if row < 0:
            Trace.error(f"sheet '{sheet_name}' - column '{column_name}' not found")
            continue

        for i in range(3, ws.max_row + 1):
            row_cells = ws[i]

            ws.cell(i, row + 1).value = ""

            if get_cell_text(row_cells[0]) != "":
                if str(i) in data_info_sheet:
                    ws.cell(i, row + 1).value = data_info_sheet[str(i)]
                else:
                    ws.cell(i, row + 1).value = 0

                if get_cell_text(row_cells[1]) != "":
                    ws.cell(i, row + 1).style = "used"

    dest_path = pathname / filename_update
    try:
        wb.save(filename=dest_path)
        Trace.result(f"'{dest_path}'")
        return True
    except OSError as err:
        Trace.error(f"{err}")
        return False
